from ast import If
from logging import _srcfile
from multiprocessing import Value
from operator import index
import string
from openpyxl import load_workbook
from utils.valve import Valve
from utils.valve2 import Valve2
from utils.nozzle import Nozzle
from utils.line import Line
from utils.valve3 import Valve3
from utils.split import Split
from utils.pump import Pump
from utils.tankreturn import TankReturn
from utils.pit import Pit
from tkinter import messagebox
import pandas as pd

###
###file for regular use: '//hanford/data/sitedata/WasteTransferEng/Waste Transfer Engineering/1 Transfers/1C - Procedure Review Tools/MasterProcedureData.xlsx'
### TEST FILE !!!change back file to regular use file when complete !!!

def importComponents(filepath='//hanford/data/sitedata/WasteTransferEng/Waste Transfer Engineering/2 Team Members/Sarah Hunter/transfer-pro - related files/DONT USE MasterProcedureData 2025-03-19.xlsx'):
    try:
        wb = load_workbook(filename=filepath, data_only=True)
    except FileNotFoundError as e:
        raise e

    pits_sheet = wb["Pit Components"]
    pits = {}

    for row in pits_sheet.iter_rows(min_row=3, values_only=True, max_row=150, max_col=26):
        pit_name = row[1]
        # if a pit hasn't been initialized as an object, create it
        if pit_name not in pits:
            pits[pit_name] = Pit(
                name=pit_name, leak_detector_id=row[2], leak_detector_pmid=row[3],
                leak_detector_tfsps=row[4], tfsps_transmitter=row[5],
                tfsps_pmid=row[6], drain_seal_location=row[7], drain_seal_name=[8],
                drain_seal_position=row[9], annulus_leak_detector=row[12],
                annulus_leak_detector_pmid=row[13], pit_nace=row[14],
                pit_nace_pmid=row[15], in_pit_heater=row[16],
                tfmcs=row[17], tsr_structure=row[18]
            )
        # if a pit has been initialized, add the other fields in the excel row
        else:
            pits[pit_name].add_missing_fields(
                tfsps_transmitter=row[5], tfsps_pmid=row[6], 
                annulus_leak_detector=row[12], annulus_leak_detector_pmid=row[13],
                in_pit_heater=row[16],
            )

    # "Components" dictionary to store and retrieve all objects representing waste transffer valves, pumps, lines, etc.
    components = { }
    
    # Types dictionary to initialize objects of the Class specified in the Excel sheet for each component 
    types = {
        "2-Way-Valve": Valve2,
        "3-Way-Valve": Valve3,
        "Tee Fitting": Split,
        "Split Point": Split,
        "Pump": Pump,
        "Transfer Line": Line,
        "Nozzle": Nozzle,
        "": Valve,
        "Tank Return": TankReturn,
        None: Valve
    }

    # Dictionary for virtual keys that were created to establish a relationship between each EIN and Connection1, Connection2, Connection3.
    #??? Probably don't need anymore ?
    connections_set_id = { }
    

    connections_sheet = wb['Transfer Route Components']
    connections_matrix = connections_sheet["F2:H400"]
    
    #Create existing data structure and new data stucture dataframes so data manipulation is easier. 
    #Cnx is an abbreviation for connections.
    dfOriginal = pd.read_excel(filepath, "Transfer Route Components",  index_col=None, na_values='NULL')
    dfCnx = dfOriginal[["Connection Set ID", "Component", "Connection 1", "Connection 2", "Connection 3", 
                                                                             "Connection Label 1", "Connection Label 2", "Connection Label 3"]]
    dfCnx1 = pd.melt(dfCnx, id_vars = ['Component'], 
                         value_vars = ['Connection 1', 'Connection 2', 'Connection 3'], 
                         var_name = 'Cnx #', 
                         value_name = 'Cnx')
    dfCnxLabels = pd.melt(dfCnx, id_vars = ['Component'], 
                         value_vars = ['Connection Label 1', 'Connection Label 2', 'Connection Label 3'], 
                         var_name = 'Label #', 
                         value_name = 'Cnx Label')
    #Get the connection number from the variable names
    dfCnx1['Cnx #'] = dfCnx1['Cnx #'].str.extract('(\d)').astype(int) 
    dfCnxLabels['Label #'] = dfCnxLabels['Label #'].str.extract('(\d)').astype(int)
    #Join the connections and label tables and put them in order
    dfCnx_result = pd.merge(dfCnx1, dfCnxLabels, left_on=['Component', 'Cnx #'], right_on=['Component', 'Label #'])
    dfCnx_result = dfCnx_result.drop(columns=['Label #'])
    dfCnx_result = dfCnx_result.sort_values(by=['Component', 'Cnx #']).reset_index(drop=True)
    dfCnx_result['Cnx Key'] = range(1, len(dfCnx_result) + 1)
    print(dfCnx_result)
    #dfCnx_result.to_excel('dfCnx_result.xlsx', index=False)


    #Initialize component objects into "components" dictionary. Key = Name, value = Component object
    for row in connections_sheet.iter_rows(min_row=2, values_only= True, max_row=350, max_col=15):
        component_name = row[0]
        component_type = row[1]
        components[component_name] = types[component_type](component_name, pit = row[2], jumper = row[3], dvi = row[4], field_label = row[12])
        
 

    #For all component object in "components". Call their .conect() method to its neighbors
    for row_index1, (component, neighbors) in enumerate(zip(components.values(), connections_matrix)):
        for col_index1, neighbor in enumerate(neighbors):
            #if the neighbor listed in excel exists as an initialized object in the "components" dictionary
            if neighbor.value in components:
                #if neighbor object != None (might be a reduntant check)
                if neighbor.value:
                    component.connect(components[neighbor.value])
                    #get cnx key of each neighbor and make a dataframe or dictionary?


    return components, pits, connections_set_id
importComponents()
